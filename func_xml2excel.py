from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
import re
import os
from rtde.util import return_print,set_log,create_engine
import copy
import functools

def get_xml_file_list(path):
    src_dir_name = ''
    with open('./xml_file_list.txt',mode='r',encoding='utf-8') as xml_file:
        xml_file_list = [path+xml.strip('\n') for xml in xml_file.readlines() if os.path.isfile(path+xml.strip('\n'))]
    return xml_file_list

print(get_xml_file_list('D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/'))

def get_group_exec_ord_no():
    try:
        engine = create_engine(args=None)
        with engine.connect() as conn:
            sql = "select max(exec_ord_no)+1 as exec_ord_no from ais8801 where aval_ed_dt ='99991231235959' and lia_type_cd ='0'"
            result = conn.execute(sql)
            return [row["exec_ord_no"] for row in result][0]
    except Exception as e:
        print(e)

def get_soup(xml_path:str):
    try:
        data_file = open(xml_path,'r', encoding='utf-8')
        soup = BeautifulSoup(data_file,'xml')
    except Exception as e:
        print(e)
    finally:
        data_file.close()
    return soup


print(get_group_exec_ord_no())

def make_group_excel_sheet(xml_path,sheet_name,sheet_order,excel_path,exec_ord_no,program_group_id,date_string):
    write_wb = Workbook()

    # sheet생성
    write_group = write_wb.create_sheet(sheet_name,sheet_order)

    #헤더지정
    write_group['A1'] = "프로그램그룹ID"
    write_group['B1'] = "유효종료일자"
    write_group['C1'] = "유효시작일자"
    write_group['D1'] = "프로그램그룹명"
    write_group['E1'] = "프로그램그룹설명"
    write_group['F1'] = "LIA유형코드"
    write_group['G1'] = "실행순번"
    write_group['H1'] = "재분석실행여부"
    write_group['I1'] = "실행조건여부"
    write_group['J1'] = "실행조건SQLID"
    write_group['K1'] = "변경일시"
    write_group['L1'] = "변경자"

    soup = get_soup(xml_path)

    ## mapper 태그
    tag = soup.find("mapper")
    namespace = tag.attrs.get('namespace')

    write_group.cell(2,1,program_group_id)
    write_group.cell(2,2,'99991231235959')
    write_group.cell(2,3,date_string)
    write_group.cell(2,4,namespace)
    write_group.cell(2,5,' ')
    write_group.cell(2,6,'0')
    write_group.cell(2,7,exec_ord_no)
    write_group.cell(2,8,'N')
    write_group.cell(2,9,'N')
    write_group.cell(2,10,' ')
    write_group.cell(2,11,datetime.now())
    write_group.cell(2,12,'admin')
    write_wb.save(excel_path)

make_group_excel_sheet('D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/ImpactStreamInclude.xml',sheet_name='쿼리그룹',sheet_order=0,excel_path='D:/t.xlsx',exec_ord_no=43,program_group_id='impact', date_string='20230103125959')


def classify_crud(pgm_src):
    create1 = re.search(r'\sINSERT\s',pgm_src,re.I)
    create2 = re.search(r'\sCREATE\s',pgm_src,re.I)
    read = re.search(r'\sSELECT\s',pgm_src,re.I)
    update = re.search(r'\sUPDATE\s',pgm_src,re.I)
    delete1 = re.search(r'\sDELETE\s',pgm_src,re.I)
    delete2 = re.search(r'\sDROP\s',pgm_src,re.I)
    
    if create1 :
        return 'C'
    if create2 :
        return 'C'
    if read :
        return 'R' 
    if update :
        return 'U' 
    if delete1 :
        return 'D' 
    if delete2 :
        return 'D'                         

def get_program_by_tag(xml_path,tag_name):
    soup = get_soup(xml_path)
    tag = soup.find_all(tag_name)
    list = []
    if tag is None:
        return list
    for idx, e in enumerate(tag) :
        id = return_print(e.get('id'))
        pgm_id = os.path.basename(xml_path)+"."+ id
        pgm_grp_id = "com/datastreams/qtrack/mappers/"+ os.path.basename(xml_path)
        aval_ed_dt = '99991231235959' 
        aval_st_dt = datetime.now().strftime("%Y%m%d%H%M%S")
        lia_type_cd ='0'
        db_type_cd = 'postgresql'
        pgm_nm = id
        exec_ord_no = idx+1
        pgm_src = return_print(e)
        src_type_cd = '5'
        sql_type_cd = classify_crud(pgm_src)
        pgm_desc = ' '
        data_maker = '-1'
        common_exec_yn = 'N'
        use_type_cd = 'Y'
        del_yn = 'N'
        mdfy_date = datetime.now()
        mdfy_user = 'admin'
        row = (pgm_id,pgm_grp_id,aval_ed_dt,aval_st_dt,lia_type_cd,db_type_cd,pgm_nm,exec_ord_no,pgm_src,src_type_cd,sql_type_cd,pgm_desc,data_maker,common_exec_yn,use_type_cd,del_yn,mdfy_date,mdfy_user)
        list.append(row)
    return list

# print(get_program_by_tag('D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/ImpactStreamRelation.xml','insert'))

xml_path = 'D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/ImpactStreamRelation.xml'

print("*"*50)

# list = [get_program_by_tag(xml_path,tag_name) for tag_name in ['resultmap','sql','select','insert','update','delete']]

# list = filter(lambda item: True if item else False,list)

# for idx,val  in enumerate(list):
#     print(idx,val)

    


def make_program_excel_sheet(workbook,xml_path,excel_path):
    def write_excel_program_header(sheet):
        sheet['A1'] = "프로그램ID"
        sheet['B1'] = "프로그램그룹ID"
        sheet['C1'] = "유효종료일자"
        sheet['D1'] = "유효시작일자"
        sheet['E1'] = "LIA유형코드"
        sheet['F1'] = "DB유형코드"
        sheet['G1'] = "프로그램명"
        sheet['H1'] = "프로그램실행순번"
        sheet['I1'] = "소스내용"
        sheet['J1'] = "소스유형코드"
        sheet['K1'] = "SQL유형코드"
        sheet['L1'] = "프로그램설명"
        sheet['M1'] = "데이터생성구분"
        sheet['N1'] = "공통실행여부"
        sheet['O1'] = "사용유형코드"
        sheet['P1'] = "삭제여부"
        sheet['Q1'] = "변경일시"
        sheet['R1'] = "변경자"
        return sheet
    def write_excel_program(list, sheet):
        copy_sheet = sheet
        copy_list = list
        print(copy_list)
        for val in enumerate(copy_list):
            idx = 2
            for inner_idx,inner_list in enumerate(val):
                print(inner_list)
                # copy_sheet.cell(idx,1,inner_list[inner_idx][0])
                # copy_sheet.cell(idx,2,inner_list[inner_idx][1])
                # copy_sheet.cell(idx,3,inner_list[inner_idx][2])
                # copy_sheet.cell(idx,4,inner_list[inner_idx][3])
                # copy_sheet.cell(idx,5,inner_list[inner_idx][4])
                # copy_sheet.cell(idx,6,inner_list[inner_idx][5])
                # copy_sheet.cell(idx,7,inner_list[inner_idx][6])
                # copy_sheet.cell(idx,8,inner_list[inner_idx][7])
                # copy_sheet.cell(idx,9,inner_list[inner_idx][8])
                # copy_sheet.cell(idx,10,inner_list[inner_idx][9])
                # copy_sheet.cell(idx,11,inner_list[inner_idx][10])
                # copy_sheet.cell(idx,12,inner_list[inner_idx][11])
                # copy_sheet.cell(idx,13,inner_list[inner_idx][12])
                # copy_sheet.cell(idx,14,inner_list[inner_idx][13])
                # copy_sheet.cell(idx,15,inner_list[inner_idx][14])
                # copy_sheet.cell(idx,16,inner_list[inner_idx][15])
                # copy_sheet.cell(idx,17,inner_list[inner_idx][16])
                # copy_sheet.cell(idx,18,inner_list[inner_idx][17])
            idx = idx + 1
        return copy_sheet
    
    write_wb = copy.deepcopy(workbook)
    write_progam = write_wb.create_sheet('쿼리',1)
    write_progam = write_excel_program_header(write_progam)

    list = [get_program_by_tag(xml_path,tag_name) for tag_name in ['resultmap','sql','select','insert','update','delete']]
    list = filter(lambda item: True if item else False,list)

    write_progam = write_excel_program(list, write_progam)
    write_wb.save(excel_path)

print('시작'*50)
workbook = Workbook()
excel_path = 'D:/test.xlsx'
make_program_excel_sheet(workbook,xml_path,excel_path)
print('종료'*50)
exit()
print('시작')

xml_list = ['ImpactStreamInclude','ImpactStreamCommon','ImpactStreamInfo','ImpactStreamRelation','ImpactStreamFileDb','ImpactStreamHistory']
#xml_list = ['ImpactStreamInfo']
start_exec_ord_no = 37

array_index = 0
for xml in xml_list :
    # 상수정의
    xml_path = 'D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/'+xml_list[array_index]+'.xml'
    excel_path =(os.path.expanduser('~')+'\\downloads').replace('\\','/')+'/RTDE_'+xml_list[array_index]+'.xlsx'
    print('excel_path:'+excel_path)
    program_group_id ='com/datastreams/qtrack/mappers/' + xml_list[array_index]+ '.xml'
    programPrefix = xml_list[array_index]+'.xml.'

    #날짜 지정
    dt_string = datetime.now().strftime("%Y%m%d%H%M%S")

    write_wb = Workbook()

    # sheet생성
    write_group = write_wb.create_sheet('쿼리그룹',0)

    #헤더지정
    write_group['A1'] = "프로그램그룹ID"
    write_group['B1'] = "유효종료일자"
    write_group['C1'] = "유효시작일자"
    write_group['D1'] = "프로그램그룹명"
    write_group['E1'] = "프로그램그룹설명"
    write_group['F1'] = "LIA유형코드"
    write_group['G1'] = "실행순번"
    write_group['H1'] = "재분석실행여부"
    write_group['I1'] = "실행조건여부"
    write_group['J1'] = "실행조건SQLID"
    write_group['K1'] = "변경일시"
    write_group['L1'] = "변경자"

    with open(xml_path,'r', encoding='utf-8') as data_file:
        soup = BeautifulSoup(data_file,'xml')

        ## mapper 태그
        tag = soup.find("mapper")
        namespace = tag.attrs.get('namespace')

        write_group.cell(2,1,program_group_id)
        write_group.cell(2,2,'99991231235959')
        write_group.cell(2,3,dt_string)
        write_group.cell(2,4,namespace)
        write_group.cell(2,5,' ')
        write_group.cell(2,6,'0')
        write_group.cell(2,7,start_exec_ord_no)
        write_group.cell(2,8,'N')
        write_group.cell(2,9,'N')
        write_group.cell(2,10,' ')
        write_group.cell(2,11,datetime.now())
        write_group.cell(2,12,'admin')
        write_wb.save(excel_path)

        start_exec_ord_no = start_exec_ord_no +1


        write_progam = write_wb.create_sheet('쿼리',1)
        write_progam['A1'] = "프로그램ID"
        write_progam['B1'] = "프로그램그룹ID"
        write_progam['C1'] = "유효종료일자"
        write_progam['D1'] = "유효시작일자"
        write_progam['E1'] = "LIA유형코드"
        write_progam['F1'] = "DB유형코드"
        write_progam['G1'] = "프로그램명"
        write_progam['H1'] = "프로그램실행순번"
        write_progam['I1'] = "소스내용"
        write_progam['J1'] = "소스유형코드"
        write_progam['K1'] = "SQL유형코드"
        write_progam['L1'] = "프로그램설명"
        write_progam['M1'] = "데이터생성구분"
        write_progam['N1'] = "공통실행여부"
        write_progam['O1'] = "사용유형코드"
        write_progam['P1'] = "삭제여부"
        write_progam['Q1'] = "변경일시"
        write_progam['R1'] = "변경자"

        ## result map 태그
        tag = soup.find_all("resultMap")

        i = 2
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## sql 태그
        tag = soup.find_all("sql")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1   

        ## select 태그
        tag = soup.find_all("select")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            
            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## insert 태그
        tag = soup.find_all("insert")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
                    
            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'I'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## update 태그
        tag = soup.find_all("update")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1

            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            
            m1 = re.search(r'\sCREATE\s',e_10,re.I)
            m2 = re.search(r'\sDROP\s',e_10,re.I)
            if m1 :
                e_11 = 'C'
            elif m2 :
                e_11 = 'D'
            else :
                e_11 = 'U' 

            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1 

        ## delete 태그
        tag = soup.find_all("delete")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1

            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'D'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1  
    array_index = array_index +1
write_wb.close()
print('종료')

if __name__ == "__main__":
    print('함수형 xml2excel')