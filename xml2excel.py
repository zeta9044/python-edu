from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
import re
import os
from rtde.util import return_print

print('시작')

xml_list = ['ImpactStreamInclude','ImpactStreamCommon','ImpactStreamInfo','ImpactStreamRelation','ImpactStreamFileDb','ImpactStreamHistory']
#xml_list = ['ImpactStreamInfo']
start_exec_ord_no = 37

array_index = 0
for xml in xml_list :
    # 상수정의
    xml_path = 'D:/git/qtrack/qtrack-server/src/main/resources/com/datastreams/qtrack/mappers/postgresql/'+xml_list[array_index]+'.xml'
    excel_path =(os.path.expanduser('~')+'\\downloads').replace('\\','/')+'/RTDE_'+xml_list[array_index]+'.xlsx'
    print('excel_path:'+excel_path)
    program_group_id ='com/datastreams/qtrack/mappers/' + xml_list[array_index]+ '.xml'
    programPrefix = xml_list[array_index]+'.xml.'

    #날짜 지정
    dt_string = datetime.now().strftime("%Y%m%d%H%M%S")

    write_wb = Workbook()

    # sheet생성
    write_group = write_wb.create_sheet('쿼리그룹',0)

    #헤더지정
    write_group['A1'] = "프로그램그룹ID"
    write_group['B1'] = "유효종료일자"
    write_group['C1'] = "유효시작일자"
    write_group['D1'] = "프로그램그룹명"
    write_group['E1'] = "프로그램그룹설명"
    write_group['F1'] = "LIA유형코드"
    write_group['G1'] = "실행순번"
    write_group['H1'] = "재분석실행여부"
    write_group['I1'] = "실행조건여부"
    write_group['J1'] = "실행조건SQLID"
    write_group['K1'] = "변경일시"
    write_group['L1'] = "변경자"

    with open(xml_path,'r', encoding='utf-8') as data_file:
        soup = BeautifulSoup(data_file,'xml')

        ## mapper 태그
        tag = soup.find("mapper")
        namespace = tag.attrs.get('namespace')

        write_group.cell(2,1,program_group_id)
        write_group.cell(2,2,'99991231235959')
        write_group.cell(2,3,dt_string)
        write_group.cell(2,4,namespace)
        write_group.cell(2,5,' ')
        write_group.cell(2,6,'0')
        write_group.cell(2,7,start_exec_ord_no)
        write_group.cell(2,8,'N')
        write_group.cell(2,9,'N')
        write_group.cell(2,10,' ')
        write_group.cell(2,11,datetime.now())
        write_group.cell(2,12,'admin')
        write_wb.save(excel_path)

        start_exec_ord_no = start_exec_ord_no +1


        write_progam = write_wb.create_sheet('쿼리',1)
        write_progam['A1'] = "프로그램ID"
        write_progam['B1'] = "프로그램그룹ID"
        write_progam['C1'] = "유효종료일자"
        write_progam['D1'] = "유효시작일자"
        write_progam['E1'] = "LIA유형코드"
        write_progam['F1'] = "DB유형코드"
        write_progam['G1'] = "프로그램명"
        write_progam['H1'] = "프로그램실행순번"
        write_progam['I1'] = "소스내용"
        write_progam['J1'] = "소스유형코드"
        write_progam['K1'] = "SQL유형코드"
        write_progam['L1'] = "프로그램설명"
        write_progam['M1'] = "데이터생성구분"
        write_progam['N1'] = "공통실행여부"
        write_progam['O1'] = "사용유형코드"
        write_progam['P1'] = "삭제여부"
        write_progam['Q1'] = "변경일시"
        write_progam['R1'] = "변경자"

        ## result map 태그
        tag = soup.find_all("resultMap")

        i = 2
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## sql 태그
        tag = soup.find_all("sql")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1   

        ## select 태그
        tag = soup.find_all("select")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
            
            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'R'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## insert 태그
        tag = soup.find_all("insert")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1
                    
            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'I'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1

        ## update 태그
        tag = soup.find_all("update")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1

            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            
            m1 = re.search(r'\sCREATE\s',e_10)
            m2 = re.search(r'\sDROP\s',e_10)
            if m1 :
                e_11 = 'C'
            elif m2 :
                e_11 = 'D'
            else :
                e_11 = 'U' 

            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1 

        ## delete 태그
        tag = soup.find_all("delete")

        i = i
        for e in tag :
            id = return_print(e.get('id'))
            e_01 = programPrefix + id
            e_02 = program_group_id
            e_03 = '99991231235959' 
            e_04 = dt_string
            e_05 ='0'
            e_06 = 'postgresql'
            e_07 = id
            e_08 = i-1

            #쿼리 앞에 ID 붙이기
            change_string = '/* ' + e_01 + ' */ '
            e.contents[0].string.insert_before(change_string)

            e_09 = return_print(e)
            e_10 = '5'
            e_11 = 'D'
            e_12 = ' '
            e_13 = '-1'
            e_14 = 'N'
            e_15 = 'Y'
            e_16 = 'N'
            e_17 = datetime.now()
            e_18 = 'admin'
            write_progam.cell(i,1,e_01)
            write_progam.cell(i,2,e_02)
            write_progam.cell(i,3,e_03)
            write_progam.cell(i,4,e_04)
            write_progam.cell(i,5,e_05)
            write_progam.cell(i,6,e_06)
            write_progam.cell(i,7,e_07)
            write_progam.cell(i,8,e_08)
            write_progam.cell(i,9,e_09)
            write_progam.cell(i,10,e_10)
            write_progam.cell(i,11,e_11)
            write_progam.cell(i,12,e_12)
            write_progam.cell(i,13,e_13)
            write_progam.cell(i,14,e_14)
            write_progam.cell(i,15,e_15)
            write_progam.cell(i,16,e_16)
            write_progam.cell(i,17,e_17)
            write_progam.cell(i,18,e_18)
            write_wb.save(excel_path)
            i = i+1  
    array_index = array_index +1
write_wb.close()
print('종료')